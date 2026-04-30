"""
文档解析器模块
支持多种文档格式的解析
"""

import os
from typing import Optional
from abc import ABC, abstractmethod


class BaseParser(ABC):
    """解析器基类"""
    
    @abstractmethod
    def parse(self, file_path: str) -> str:
        """解析文档，返回文本内容"""
        pass


class PDFParser(BaseParser):
    """PDF 文档解析器"""
    
    def parse(self, file_path: str) -> str:
        """解析 PDF 文件"""
        try:
            import PyPDF2
            
            with open(file_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                text = ""
                for page in reader.pages:
                    text += page.extract_text() + "\n"
                return text
        except Exception as e:
            raise Exception(f"PDF 解析失败: {str(e)}")


class DocxParser(BaseParser):
    """Word 文档解析器"""
    
    def parse(self, file_path: str) -> str:
        """解析 DOCX 文件"""
        try:
            from docx import Document
            
            doc = Document(file_path)
            text = ""
            for para in doc.paragraphs:
                text += para.text + "\n"
            return text
        except Exception as e:
            raise Exception(f"DOCX 解析失败: {str(e)}")


class ExcelParser(BaseParser):
    """Excel 文档解析器"""
    
    def parse(self, file_path: str) -> str:
        """解析 XLSX 文件"""
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(file_path)
            text = ""
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text += f"=== Sheet: {sheet_name} ===\n"
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    text += row_text + "\n"
                text += "\n"
            
            return text
        except Exception as e:
            raise Exception(f"XLSX 解析失败: {str(e)}")


class TxtParser(BaseParser):
    """纯文本文档解析器"""
    
    def parse(self, file_path: str) -> str:
        """解析 TXT 文件"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                return file.read()
        except UnicodeDecodeError:
            # 尝试其他编码
            with open(file_path, 'r', encoding='gbk') as file:
                return file.read()
        except Exception as e:
            raise Exception(f"TXT 解析失败: {str(e)}")


class DocumentParser:
    """统一文档解析器"""
    
    def __init__(self):
        self.parsers = {
            '.pdf': PDFParser(),
            '.docx': DocxParser(),
            '.xlsx': ExcelParser(),
            '.xls': ExcelParser(),
            '.txt': TxtParser(),
            '.md': TxtParser()
        }
    
    def parse(self, file_path: str) -> str:
        """解析文档
        
        Args:
            file_path: 文档路径
            
        Returns:
            文档文本内容
        """
        ext = os.path.splitext(file_path)[1].lower()
        
        if ext not in self.parsers:
            raise ValueError(f"不支持的文件格式: {ext}")
        
        parser = self.parsers[ext]
        return parser.parse(file_path)
    
    def get_supported_formats(self) -> list:
        """获取支持的文档格式"""
        return list(self.parsers.keys())
